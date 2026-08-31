"""Pure, generic export builders for the config-driven report portal (L4).

Turns a report's ordered ``ColumnSpec`` list + the same row dicts the table
renders into CSV or XLSX bytes, with the acknowledged data-handling disclaimer
riding at the TOP of every file. The disclaimer is a single source of truth: the
same resolved text is the acknowledgement checkbox label in the template AND the
text embedded in each export. ``DEFAULT_DISCLAIMER`` is the built-in fallback;
the effective text is resolved from the ``DOWNLOAD_DISCLAIMER`` env var.

This module is stdlib-only at import time. ``openpyxl`` is imported lazily
INSIDE :func:`to_xlsx_bytes`, so the pytest-only dev ``.venv`` (no openpyxl) can
still import this module and run the CSV/audit tests.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Iterable, TextIO

# Dual-convention imports: the Apps runtime runs from ``src/app/`` with flat
# imports; the pytest suite imports this module as ``app.exports`` with ``src``
# on the path. Try the flat form first, fall back to the package form.
try:  # pragma: no cover - import shim
    from render import cell_text
    from reports import ColumnSpec
    from shaping import _to_int
except ImportError:  # pragma: no cover - import shim
    from app.render import cell_text
    from app.reports import ColumnSpec
    from app.shaping import _to_int

# Generic default acknowledgement text (checkbox label in the download modal AND
# the text embedded at the top of every export). Override per-deployment with the
# DOWNLOAD_DISCLAIMER env var (resolved in config.resolve_disclaimer); this
# constant is the fallback so the app always has a sensible notice.
DEFAULT_DISCLAIMER: str = (
    "DATA HANDLING NOTICE.\n"
    "This export contains data provided for authorized use only.\n"
    "By downloading this file you acknowledge that: (1) you are authorized to access "
    "this data; (2) you will handle it in accordance with your organization's data-"
    "handling and safeguarding policy; (3) you will not redistribute, publish, or "
    "share it with unauthorized parties; and (4) you will store and dispose of it "
    "securely.\n"
    "This download is logged and attributed to your identity."
)


def filename_for(report_id: str, date: str, fmt: str) -> str:
    """Build the attachment filename for an export.

    Args:
        report_id: The report registry key (e.g. ``daily_metrics``).
        date: The selected report date (``"%Y-%m-%d %H:%M:%S"`` or a bare date);
            the date portion is used (time stripped).
        fmt: The export format, ``"csv"`` or ``"xlsx"``.

    Returns:
        A filename like ``daily_metrics_2026-01-12.csv``.
    """
    date_part = (date or "").strip().split(" ")[0]
    ext = "xlsx" if fmt == "xlsx" else "csv"
    return f"{report_id}_{date_part}.{ext}"


# Characters that would break or inject a quoted Content-Disposition filename:
# the double-quote (closes the quoted-string), the backslash (escape), and any
# ASCII control char incl. CR/LF (header splitting).
_UNSAFE_FILENAME_CHARS = re.compile(r'[\x00-\x1f"\\]')


def sanitize_filename(name: str) -> str:
    """Sanitize a filename for safe use in a ``Content-Disposition`` header.

    Volume file names come from arbitrary Unity Catalog volume contents, so a
    name containing a double-quote or CR/LF could truncate the header or smuggle
    a second header. Replace every unsafe character (``"``, ``\\``, and ASCII
    control chars including CR/LF) with ``_`` and trim surrounding whitespace.

    Args:
        name: The raw filename (e.g. a volume file basename).

    Returns:
        A header-safe filename; ``"download"`` if nothing usable remains.
    """
    cleaned = _UNSAFE_FILENAME_CHARS.sub("_", name or "").strip()
    return cleaned or "download"


def to_csv_bytes(
    columns: list[ColumnSpec], rows: list[dict], disclaimer: str
) -> bytes:
    """Render the export rows as CSV bytes with the disclaimer at the top.

    When ``disclaimer`` has content it rides at the very top as leading
    ``# ``-prefixed single-cell rows, then a blank separator row, then the label
    header row, then the data rows. When ``disclaimer`` is empty/blank (e.g. the
    admin audit-log export) neither the disclaimer rows nor the blank separator
    are emitted, so the header is the first row. Every cell is formatted via
    :func:`render.cell_text` (all strings), so the CSV matches the on-screen
    table exactly.

    Args:
        columns: The report's ordered display columns.
        rows: The snapshot row dicts keyed by column name.
        disclaimer: The data-handling text to embed at the top (normally
            :data:`DISCLAIMER`).

    Returns:
        UTF-8 encoded CSV bytes.
    """
    buf = io.StringIO()
    write_csv(buf, columns, rows, disclaimer)
    return buf.getvalue().encode("utf-8")


def write_csv(
    file: TextIO,
    columns: list[ColumnSpec],
    rows: Iterable[dict],
    disclaimer: str = "",
    *,
    include_header: bool = True,
) -> None:
    """Write CSV rows incrementally without building the export in memory.

    ``include_header=False`` supports appending later query pages to the same
    file. This is the volume-delivery path for large exports.
    """
    writer = csv.writer(file)
    # Only emit the disclaimer block + blank separator when there is disclaimer
    # text; an empty/blank disclaimer (e.g. the admin audit-log export) must NOT
    # produce a leading blank row before the header, or parsers see a blank
    # header and misaligned columns.
    if include_header and disclaimer.strip():
        for line in disclaimer.splitlines():
            writer.writerow([f"# {line}"])
        writer.writerow([])  # blank separator row
    if include_header:
        writer.writerow([c.label for c in columns])
    for row in rows:
        writer.writerow([cell_text(row.get(c.name), c.format) for c in columns])


def to_xlsx_bytes(
    columns: list[ColumnSpec], rows: list[dict], disclaimer: str
) -> bytes:
    """Render the export rows as XLSX bytes with the disclaimer at the top.

    ``openpyxl`` is imported lazily inside this function so the module stays
    stdlib-only at import time. The disclaimer occupies merged, wrapped, italic
    top rows, then a blank spacer row, then a bold label header, then the data.
    Per column: ``int`` columns are written as real numeric cells
    (``_to_int``, so Excel treats them as sortable/summable numbers); ``pct`` and
    ``text``/unknown columns are written as their display string via
    :func:`render.cell_text` (so ``—`` and the signed ``%`` match the screen).

    Args:
        columns: The report's ordered display columns.
        rows: The snapshot row dicts keyed by column name.
        disclaimer: The data-handling text to embed at the top (normally
            :data:`DISCLAIMER`).

    Returns:
        XLSX file bytes (a ZIP container beginning with ``b"PK"``).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    ncols = max(1, len(columns))
    wb = Workbook()
    ws = wb.active

    r = 1
    for line in disclaimer.splitlines():
        ws.cell(row=r, column=1, value=line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(italic=True)
        r += 1

    r += 1  # blank spacer row

    for col, c in enumerate(columns, start=1):
        header_cell = ws.cell(row=r, column=col, value=c.label)
        header_cell.font = Font(bold=True)
    r += 1

    for row in rows:
        for col, c in enumerate(columns, start=1):
            if c.format == "int":
                ws.cell(row=r, column=col, value=_to_int(row.get(c.name)))
            else:
                ws.cell(row=r, column=col, value=cell_text(row.get(c.name), c.format))
        r += 1

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
